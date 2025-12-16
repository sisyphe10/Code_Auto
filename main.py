import time
import schedule
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
import os
import yfinance as yf
import re

# === 상수 정의 ===
# [수정됨] GitHub 저장소에 바로 저장되도록 경로를 파일명으로만 변경했습니다.
EXCEL_FILE = 'dataset.xlsx'

# DRAM 제품명
TARGET_DRAM_ITEMS = {
    'DDR5 16G (2Gx8) 4800/5600': 'DDR5 16G (2Gx8) 4800/5600',
    'DDR4 16Gb (1Gx16)3200': 'DDR4 16Gb (1Gx16)3200',
    'DDR4 16Gb (2Gx8)3200': 'DDR4 16Gb (2Gx8)3200',
    'DDR4 8Gb (1Gx8) 3200': 'DDR4 8Gb (1Gx8) 3200',
    'DDR4 8Gb (512Mx16) 3200': 'DDR4 8Gb (512Mx16) 3200'
}

# NAND 제품명
TARGET_NAND_ITEMS = {
    'SLC 2Gb 256MBx8': 'SLC 2Gb 256MBx8',
    'SLC 1Gb 128MBx8': 'SLC 1Gb 128MBx8',
    'MLC 64Gb 8GBx8': 'MLC 64Gb 8GBx8',
    'MLC 32Gb 4GBx8': 'MLC 32Gb 4GBx8'
}

# yfinance 및 금리 티커 목록
YFINANCE_TICKERS = {
    'Bitcoin': {'ticker': 'BTC-USD', 'type': 'CRYPTO'},
    'Ethereum': {'ticker': 'ETH-USD', 'type': 'CRYPTO'},
    'Binance Coin': {'ticker': 'BNB-USD', 'type': 'CRYPTO'},
    'Ripple': {'ticker': 'XRP-USD', 'type': 'CRYPTO'},
    'Solana': {'ticker': 'SOL-USD', 'type': 'CRYPTO'},
    'Dogecoin': {'ticker': 'DOGE-USD', 'type': 'CRYPTO'},
    'WTI Crude Oil': {'ticker': 'CL=F', 'type': 'COMMODITY'},
    'Brent Crude Oil': {'ticker': 'BZ=F', 'type': 'COMMODITY'},
    'Natural Gas': {'ticker': 'NG=F', 'type': 'COMMODITY'},
    'Gold': {'ticker': 'GC=F', 'type': 'COMMODITY'},
    'Silver': {'ticker': 'SI=F', 'type': 'COMMODITY'},
    'Copper': {'ticker': 'HG=F', 'type': 'COMMODITY'},
    'Uranium ETF (URA)': {'ticker': 'URA', 'type': 'COMMODITY'},
    'Wheat Futures': {'ticker': 'ZW=F', 'type': 'COMMODITY'},
    'VIX Index': {'ticker': '^VIX', 'type': 'INDEX'},
    'Dollar Index (DXY)': {'ticker': 'DX-Y.NYB', 'type': 'FX'},
    'KRW/USD': {'ticker': 'KRW=X', 'type': 'FX'},
    'CNY/USD': {'ticker': 'CNY=X', 'type': 'FX'},
    'EUR/USD': {'ticker': 'EURUSD=X', 'type': 'FX'},
    'TWD/USD': {'ticker': 'TWD=X', 'type': 'FX'},
    'JPY/USD': {'ticker': 'JPY=X', 'type': 'FX'},
    'US 2 Year Treasury Yield': {'ticker': '^IRX', 'type': 'INTEREST_RATE'},
    'US 10 Year Treasury Yield': {'ticker': '^TNX', 'type': 'INTEREST_RATE'},
    'US 30 Year Treasury Yield': {'ticker': '^TYX', 'type': 'INTEREST_RATE'}
}

# === 유틸리티 함수 ===
def setup_excel():
    """엑셀 파일 초기 설정"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Price Data'
        ws['A1'] = '날짜'
        ws['B1'] = '제품명'
        ws['C1'] = '가격'
        ws['D1'] = '데이터 타입'
        wb.save(EXCEL_FILE)
        print(f"✅ 엑셀 파일 생성 완료: {EXCEL_FILE}")
    else:
        print(f"✅ 기존 엑셀 파일 사용: {EXCEL_FILE}")

def setup_driver(headless=True):
    """Selenium 웹드라이버 설정"""
    chrome_options = Options()
    # GitHub Actions 등 서버 환경에서는 반드시 headless 모드를 권장합니다.
    if headless:
        chrome_options.add_argument('--headless')
    
    # 서버 환경 실행을 위한 필수 옵션들
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--window-size=1920,1080')
    chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return driver

def save_to_excel(data, sheet_name='Price Data'):
    """Excel에 데이터 저장"""
    try:
        wb = load_workbook(EXCEL_FILE)
        if sheet_name in wb.sheetnames:
             ws = wb[sheet_name]
        else:
             ws = wb.create_sheet(sheet_name)
             
        for data_date, name, price, data_type in data:
            ws.append([data_date, name, price, data_type])
        wb.save(EXCEL_FILE)
        wb.close()
        return True
    except PermissionError:
        print(f"\n❌ 엑셀 파일 저장 실패! '{EXCEL_FILE}' 파일이 열려있습니다.")
        return False
    except Exception as e:
        print(f"\n❌ 엑셀 저장 중 오류: {str(e)}")
        return False

def save_debug_html(page_source, filename):
    """디버그용 HTML 파일 저장"""
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(page_source)
        print(f"💡 페이지 HTML을 '{filename}'에 저장합니다.")
    except Exception as e:
        print(f"⚠️  HTML 저장 실패: {str(e)}")

def get_last_scfi_date():
    """엑셀에서 마지막으로 저장된 SCFI 날짜 가져오기"""
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        last_date = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[3] == 'OCEAN_FREIGHT' and row[0] and 'SCFI' in str(row[1]):
                last_date = row[0]
        wb.close()
        return last_date
    except Exception as e:
        print(f"⚠️  마지막 SCFI 날짜 확인 중 오류: {str(e)}")
        return None

# === 크롤링 함수 ===
def crawl_dram_nand(data_type, debug_mode=False):
    """DRAM 및 NAND 가격 크롤링"""
    print(f"\n{'=' * 60}")
    print(f"📊 {data_type} 가격 크롤링 시작")
    print(f"{'=' * 60}")

    driver = None
    success = False

    try:
        # GitHub Actions에서는 항상 headless로 실행되도록 기본값 유지
        driver = setup_driver(headless=not debug_mode)
        url = f'https://www.dramexchange.com/#{data_type.lower()}'
        print(f"🌐 웹사이트 접속 중: {url}")
        driver.get(url)

        print("⏳ 페이지 로딩 중... (최대 10초 대기)")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.TAG_NAME, 'table'))
        )

        now = datetime.now()
        current_date = now.strftime('%Y-%m-%d')
        print(f"\n📊 {data_type} Spot Price 테이블 검색 중...")

        collected_data = []
        target_items = TARGET_DRAM_ITEMS if data_type == 'DRAM' else TARGET_NAND_ITEMS
        tables = driver.find_elements(By.TAG_NAME, 'table')

        for table in tables:
            rows = table.find_elements(By.TAG_NAME, 'tr')
            for row in rows:
                cells = row.find_elements(By.TAG_NAME, 'td')
                if not cells:
                    cells = row.find_elements(By.TAG_NAME, 'th')
                if len(cells) < 2:
                    continue

                item_name = cells[0].text.strip()
                if item_name in target_items:
                    try:
                        daily_high = cells[1].text.strip()
                        if daily_high and daily_high.replace('.', '').replace(',', '').isdigit():
                            price_numeric = float(daily_high.replace(',', ''))
                            collected_data.append((current_date, item_name, price_numeric, data_type))
                            print(f"✓ {item_name}: ${daily_high}")
                    except Exception as e:
                        if debug_mode:
                            print(f"  ⚠️  {item_name} 가격 추출 실패: {str(e)}")

        if collected_data:
            if save_to_excel(collected_data):
                print(f"\n✅ {data_type}: {len(collected_data)}개 항목 저장 완료!")
                success = True
        else:
            print(f"\n⚠️  {data_type} 데이터를 수집하지 못했습니다.")

        missing_items = set(target_items.keys()) - set([item[1] for item in collected_data])
        if missing_items and debug_mode:
            print(f"\n⚠️  다음 {data_type} 항목을 찾지 못했습니다:")
            for item in missing_items:
                print(f"   - {item}")

    except Exception as e:
        print(f"\n❌ {data_type} 크롤링 오류: {str(e)}")
        if debug_mode:
            save_debug_html(driver.page_source, f"{data_type.lower()}_page_source.html")
            import traceback
            print(traceback.format_exc())

    finally:
        if driver:
            driver.quit()

    return success

def crawl_scfi_index(debug_mode=False):
    """SCFI 지수 크롤링"""
    print(f"\n{'=' * 60}")
    print(f"🚢 SCFI 지수 크롤링 시작")
    print(f"{'=' * 60}")

    driver = None
    success = False

    try:
        driver = setup_driver(headless=not debug_mode)
        url = 'https://en.sse.net.cn/indices/scfinew.jsp'
        print(f"🌐 웹사이트 접속 중: {url}")
        driver.get(url)

        print("⏳ 페이지 로딩 중... (최대 10초 대기)")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, 'currdate'))
        )

        print("\n📊 SCFI Comprehensive Index 검색 중...")
        scfi_value = None
        scfi_date = None

        try:
            date_elem = driver.find_element(By.ID, 'currdate')
            scfi_date = date_elem.text.strip()
            print(f"📅 SCFI 날짜 발견: {scfi_date}")
        except Exception as e:
            if debug_mode:
                print(f"⚠️  날짜 찾기 실패: {str(e)}")

        try:
            tables = driver.find_elements(By.TAG_NAME, 'table')
            for table in tables:
                rows = table.find_elements(By.TAG_NAME, 'tr')
                for row in rows:
                    cells = row.find_elements(By.TAG_NAME, 'td')
                    if cells and 'Comprehensive Index' in cells[0].text:
                        if debug_mode:
                            print(f"🔍 Comprehensive Index 행 발견")
                        idx4_spans = row.find_elements(By.CSS_SELECTOR, 'span.idx4')
                        if idx4_spans:
                            scfi_value = idx4_spans[0].text.strip()
                            print(f"💰 SCFI 지수 발견: {scfi_value}")
                            break
                if scfi_value:
                    break
        except Exception as e:
            if debug_mode:
                print(f"⚠️  SCFI 값 추출 실패: {str(e)}")

        if not scfi_value or not scfi_date:
            if debug_mode:
                print("\n🔍 다른 방법으로 검색 중...")
                save_debug_html(driver.page_source, 'scfi_page_source.html')

        if scfi_value and scfi_date:
            last_saved_date = get_last_scfi_date()
            if last_saved_date == scfi_date:
                print(f"\n💡 SCFI 데이터가 이미 최신입니다 (날짜: {scfi_date})")
                print(f"   새로운 업데이트가 없어 저장하지 않습니다.")
                success = True
            else:
                collected_data = [(scfi_date, 'SCFI Comprehensive Index', float(scfi_value), 'OCEAN_FREIGHT')]
                if save_to_excel(collected_data):
                    print(f"\n✅ SCFI: 새로운 데이터 저장 완료!")
                    print(f"   날짜: {scfi_date}, 지수: {scfi_value}")
                    if last_saved_date:
                        print(f"   이전 날짜: {last_saved_date}")
                    success = True
        else:
            print("\n⚠️  SCFI 데이터를 찾지 못했습니다.")
            if debug_mode:
                save_debug_html(driver.page_source, 'scfi_page_source.html')

        if debug_mode:
            # GitHub Actions에서는 입력을 받을 수 없으므로 이 부분은 패스합니다.
            pass

    except Exception as e:
        print(f"\n❌ SCFI 크롤링 오류: {str(e)}")
        if debug_mode:
            save_debug_html(driver.page_source, 'scfi_page_source.html')
            import traceback
            print(traceback.format_exc())

    finally:
        if driver:
            driver.quit()

    return success

def crawl_yfinance_data(debug_mode=False):
    """yfinance를 사용한 크립토/원자재/환율/금리 종가 크롤링"""
    print(f"\n{'=' * 60}")
    print(f"📈 yfinance 데이터 (크립토/원자재/환율/금리) 크롤링 시작")
    print(f"{'=' * 60}")

    current_date = datetime.now().strftime('%Y-%m-%d')
    collected_data = []
    success_count = 0

    try:
        for name, info in YFINANCE_TICKERS.items():
            ticker_symbol = info['ticker']
            data_type = info['type']
            try:
                ticker = yf.Ticker(ticker_symbol)
                hist = ticker.history(period='1d')
                if not hist.empty:
                    close_price = float(hist['Close'].iloc[0])
                    data_date = hist.index[0].strftime('%Y-%m-%d') if data_type != 'CRYPTO' else current_date
                    collected_data.append((data_date, name, close_price, data_type))
                    print(f"✓ {name}: {close_price:.2f}{'%' if data_type == 'INTEREST_RATE' else ''} ({data_type}, 날짜: {data_date})")
                    success_count += 1
                else:
                    print(f"⚠️  {name} 데이터 없음 (티커: {ticker_symbol})")
            except Exception as e:
                if debug_mode:
                    print(f"  ⚠️  {name} 가격 추출 실패: {str(e)}")
                else:
                    print(f"⚠️  {name} 데이터 가져오기 실패")

        if collected_data:
            if save_to_excel(collected_data):
                print(f"\n✅ yfinance 데이터: {success_count}개 항목 저장 완료!")
                return True
        else:
            print("\n⚠️  yfinance 데이터를 수집하지 못했습니다.")
            return False

    except Exception as e:
        print(f"\n❌ yfinance 크롤링 오류: {str(e)}")
        if debug_mode:
            import traceback
            print(traceback.format_exc())
        return False

def main():
    """메인 함수"""
    print("\n" + "=" * 60)
    print("🚀 전체 가격 크롤링 프로그램 (DRAM/NAND/SCFI/yfinance)")
    print("=" * 60)

    setup_excel()

    # debug_mode는 기본적으로 False로 설정하여 실행합니다.
    dram_success = crawl_dram_nand('DRAM', debug_mode=False)
    nand_success = crawl_dram_nand('NAND', debug_mode=False)
    scfi_success = crawl_scfi_index(debug_mode=False)
    yfinance_success = crawl_yfinance_data(debug_mode=False)

    print(f"\n{'=' * 60}")
    print("📊 크롤링 결과 요약")
    print(f"{'=' * 60}")
    print(f"DRAM: {'✅ 성공' if dram_success else '❌ 실패'}")
    print(f"NAND: {'✅ 성공' if nand_success else '❌ 실패'}")
    print(f"SCFI: {'✅ 성공' if scfi_success else '❌ 실패'}")
    print(f"yfinance: {'✅ 성공' if yfinance_success else '❌ 실패'}")
    print(f"{'=' * 60}")

    if any([dram_success, nand_success, scfi_success, yfinance_success]):
        print(f"\n📁 결과 파일이 {EXCEL_FILE} 에 저장되었습니다.")

if __name__ == "__main__":
    main()
